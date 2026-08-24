import tkinter as tk, os
from tkinter import ttk, messagebox
from database import get_connection, get_available_stock
from tkinter import filedialog
from tkcalendar import DateEntry
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
# from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def open_reports(win, back_command=None):
    for widget in win.winfo_children():
        widget.destroy()

    win.config(bg="#f4f6f9")

    # HEADER
    header = tk.Frame(win, bg="#f4f6f9")
    header.pack(fill="x", pady=15)

    tk.Button(header,text="← Back",command=back_command,bg="#34495e",fg="white",
            font=("Arial", 10, "bold"),width=10,height=2,relief="flat",cursor="hand2"
            ).pack(side="left", padx=20)

    tk.Label(header,text="📊 Reports",font=("Segoe UI", 22, "bold"),bg="#f4f6f9",fg="#2c3e50").pack()

    # FILTER FRAME
    filter_frame = tk.LabelFrame(win,text=" Report Filters ",font=("Segoe UI", 12, "bold"),
        bg="#f4f6f9",fg="#2c3e50",padx=15,pady=12)
    filter_frame.pack(fill="x", padx=25, pady=5)

    # Report Type
    tk.Label(filter_frame,text="Report Type:",font=("Segoe UI", 10, "bold"),
            bg="#f4f6f9").grid(row=0, column=0, padx=8, pady=5)

    report_var = tk.StringVar(value="Stock In Report")

    report_combo = ttk.Combobox(filter_frame,textvariable=report_var,state="readonly",
            width=25,font=("Segoe UI", 10))

    report_combo["values"] = (
        "Stock In Report",
        "Send Dyeing Report",
        "Return Dyeing Report",
        "Stock Out Report",
        "Supplier Report",
        "Customer Report",
        "Current Stock Report",
        "Dyeing Stock Report"    
    )

    report_combo.grid(row=0, column=1, padx=8, pady=5)

    # From Date
    tk.Label(filter_frame,text="From:",font=("Segoe UI", 10, "bold"),bg="#f4f6f9"
            ).grid(row=0, column=2, padx=8)

    from_var = tk.StringVar()
    from_entry = DateEntry(filter_frame,textvariable=from_var,width=18,
        font=("Segoe UI",10),date_pattern="yyyy-mm-dd")
    from_entry.grid(row=0,column=3,padx=8)

    tk.Label(filter_frame,text="To:",font=("Segoe UI",10,"bold"),
        bg="#f4f6f9").grid(row=0,column=4,padx=8)

    to_var = tk.StringVar()
    to_entry = DateEntry(filter_frame,textvariable=to_var,width=18,
        font=("Segoe UI",10),date_pattern="yyyy-mm-dd")
    to_entry.grid(row=0,column=5,padx=8)
    
    # BUTTONS
    button_frame = tk.Frame(win, bg="#f4f6f9")
    button_frame.pack(pady=12)

    # TABLE
    table_frame = tk.LabelFrame(win,text=" Report Data ",font=("Segoe UI", 12, "bold"),
        bg="#f4f6f9",fg="#2c3e50")
    table_frame.pack(fill="both",expand=True,padx=25,pady=5)

    tree_container = tk.Frame(table_frame,bg="#f4f6f9")
    tree_container.pack(fill="both",expand=True,padx=5,pady=5)

    tree = ttk.Treeview(tree_container,show="headings")

    tree.grid(row=0,column=0,sticky="nsew")

    v_scroll = ttk.Scrollbar(tree_container,orient="vertical",command=tree.yview)
    v_scroll.grid(row=0,column=1,sticky="ns")

    h_scroll = ttk.Scrollbar(tree_container,orient="horizontal",command=tree.xview)
    h_scroll.grid(row=1,column=0,sticky="ew")

    tree.configure(yscrollcommand=v_scroll.set,xscrollcommand=h_scroll.set)
    tree_container.grid_rowconfigure(0,weight=1)
    tree_container.grid_columnconfigure(0,weight=1)

    # SUMMARY
    summary_frame = tk.Frame(win,bg="#f4f6f9")
    summary_frame.pack(fill="x",padx=25,pady=8)

    records_label = tk.Label(summary_frame,text="Total Records: 0",font=("Segoe UI", 11, "bold"),
                bg="#f4f6f9",fg="#7e00c6")
    records_label.pack(side="left", padx=10)

    amount_label = tk.Label(summary_frame,text="Total Amount: Rs. 0",
                font=("Segoe UI", 11, "bold"),bg="#f4f6f9",fg="#00377f")
    amount_label.pack(side="right", padx=10)

    # ============================================================
    # LOAD REPORT FUNCTION
    # ============================================================
    def load_report():
        for item in tree.get_children():
            tree.delete(item)

        report = report_var.get()

        # ========== STOCK IN REPORT ==========
        if report == "Stock In Report":
            columns = ("PO #","Date","Supplier","Phone","Company","Thread","Size","Quantity","Price","Total","Paid","Balance")
            tree["columns"] = columns

            for col in columns:
                tree.heading(col,text=col)
                tree.column(col,width=100,minwidth=80,anchor="center")

            conn = None
            try:
                conn = get_connection()
                cursor = conn.cursor()

                query = """SELECT po_number,date,supplier_name,phone,company_name,
                        thread_name,size,bundle_quantity,bundle_price,paid_amount
                        FROM stock_in"""
                conditions = []
                params = []

                if from_var.get().strip():
                    conditions.append("date >= %s")
                    params.append(from_var.get().strip())

                if to_var.get().strip():
                    conditions.append("date <= %s")
                    params.append(to_var.get().strip())

                if conditions:
                    query += " WHERE " + " AND ".join(conditions)
                query += " ORDER BY id ASC"
                cursor.execute(query,params)
                records = cursor.fetchall()
                total_amount = 0

                for row in records:
                    quantity = row["bundle_quantity"] or 0
                    price = row["bundle_price"] or 0
                    paid = row["paid_amount"] or 0
                    total = quantity * price
                    balance = total - paid
                    total_amount += total

                    tree.insert("","end",values=(
                        row["po_number"],row["date"],row["supplier_name"] or "",
                        row["phone"] or "",row["company_name"] or "",
                        row["thread_name"] or "",row["size"] or "",
                        quantity,price,total,paid,balance
                    ))

                records_label.config(text=f"Total Records: {len(records)}")
                amount_label.config(text=f"Total Amount: Rs. {total_amount:,.2f}")

            except Exception as e:
                messagebox.showerror("Report Error",str(e))
            finally:
                if conn:
                    conn.close()

        # ========== SEND DYEING REPORT ==========
        elif report == "Send Dyeing Report":
            columns = ("Batch ID","Date","Thread","Size","Issued Quantity",
                    "Dyeing Info","Reason","Sender","Receiver","Status")
            tree["columns"] = columns

            for col in columns:
                tree.heading(col,text=col)
                tree.column(col,width=120,minwidth=80,anchor="center")

            conn = None
            try:
                conn = get_connection()
                cursor = conn.cursor()

                query = """SELECT batch_id,date,thread_name,size,issued_quantity,
                        dyeing_info,reason_for_issue,sender,receiver,status
                        FROM send_dyeing"""
                conditions = []
                params = []

                if from_var.get().strip():
                    conditions.append("date >= %s")
                    params.append(from_var.get().strip())

                if to_var.get().strip():
                    conditions.append("date <= %s")
                    params.append(to_var.get().strip())

                if conditions:
                    query += " WHERE " + " AND ".join(conditions)

                query += " ORDER BY id ASC"
                cursor.execute(query,params)
                records = cursor.fetchall()

                for row in records:
                    tree.insert("","end",values=(
                        row["batch_id"],row["date"],row["thread_name"] or "",
                        row["size"] or "",row["issued_quantity"] or 0,
                        row["dyeing_info"] or "",row["reason_for_issue"] or "",
                        row["sender"] or "",row["receiver"] or "",
                        row["status"] or ""
                    ))

                records_label.config(text=f"Total Records: {len(records)}")
                amount_label.config(text="Total Amount: N/A")

            except Exception as e:
                messagebox.showerror("Report Error",str(e))
            finally:
                if conn:
                    conn.close()

        # ========== RETURN DYEING REPORT ==========
        elif report == "Return Dyeing Report":
            columns = ("Date","Batch ID","Thread","Size","Color","Issued","Returned","Dyeing Info","Sender","Receiver")
            tree["columns"] = columns

            for col in columns:
                tree.heading(col,text=col)
                tree.column(col,width=110,minwidth=70,anchor="center")

            conn = None
            try:
                conn = get_connection()
                cursor = conn.cursor()

                query = """SELECT date,batch_id,thread_name,size,color,
                        issued_quantity,return_quantity,dyeing_info,sender,receiver
                        FROM return_dyeing"""
                conditions = []
                params = []

                if from_var.get().strip():
                    conditions.append("date >= %s")
                    params.append(from_var.get().strip())

                if to_var.get().strip():
                    conditions.append("date <= %s")
                    params.append(to_var.get().strip())

                if conditions:
                    query += " WHERE " + " AND ".join(conditions)

                query += " ORDER BY id ASC"
                cursor.execute(query,params)
                records = cursor.fetchall()

                for row in records:
                    tree.insert("","end",values=(
                        row["date"],
                        row["batch_id"],
                        row["thread_name"] or "",
                        row["size"] or "",
                        row["color"] or "",
                        row["issued_quantity"] or 0,
                        row["return_quantity"] or 0,
                        row["dyeing_info"] or "",
                        row["sender"] or "",
                        row["receiver"] or ""
                    ))

                records_label.config(text=f"Total Records: {len(records)}")
                amount_label.config(text="Total Amount: N/A")

            except Exception as e:
                messagebox.showerror("Report Error",str(e))
            finally:
                if conn:
                    conn.close()

        # ========== STOCK OUT REPORT ==========
        elif report == "Stock Out Report":
            columns = (
                "SO #", "Date", "Customer", "Phone", "Company",
                "Thread", "Size", "Color", "Quantity",
                "Price", "Total", "Discount", "Final Total"
            )

            tree["columns"] = columns

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=100, minwidth=70, anchor="center")
            conn = None
            try:
                conn = get_connection()
                cursor = conn.cursor()

                query = """
                    SELECT so_number, date, customer_name, phone, company_name,
                        thread_name, size, color, bundle_quantity,
                        bundle_price, discount
                    FROM stock_out
                """
                conditions = []
                params = []

                if from_var.get().strip():
                    conditions.append("date >= %s")
                    params.append(from_var.get().strip())

                if to_var.get().strip():
                    conditions.append("date <= %s")
                    params.append(to_var.get().strip())

                if conditions:
                    query += " WHERE " + " AND ".join(conditions)
                query += " ORDER BY id ASC"

                cursor.execute(query, params)
                records = cursor.fetchall()
                total_amount = 0

                for row in records:
                    quantity = row["bundle_quantity"] or 0
                    price = row["bundle_price"] or 0
                    discount = row["discount"] or 0

                    total = quantity * price
                    final_total = total - discount

                    if final_total < 0:
                        final_total = 0
                    total_amount += final_total
                    tree.insert("","end",
                        values=(
                            row["so_number"],
                            row["date"],
                            row["customer_name"] or "",
                            row["phone"] or "",
                            row["company_name"] or "",
                            row["thread_name"] or "",
                            row["size"] or "",
                            row["color"] or "",
                            quantity,
                            price,
                            total,
                            discount,
                            final_total
                        )
                    )
                records_label.config(text=f"Total Records: {len(records)}")
                amount_label.config(text=f"Total Amount: Rs. {total_amount:,.2f}")

            except Exception as e:
                messagebox.showerror("Report Error", str(e))
            finally:
                if conn:
                    conn.close()
        
        # ========== SUPPLIER REPORT ==========
        elif report == "Supplier Report":
            columns = ("Supplier","Phone","Email","CNIC","Company")
            tree["columns"] = columns

            for col in columns:
                tree.heading(col,text=col)
                tree.column(col,width=150,minwidth=80,anchor="center")

            conn = None
            try:
                conn = get_connection()
                cursor = conn.cursor()

                query = """SELECT DISTINCT supplier_name,phone,email,
                        supplier_cnic,company_name FROM stock_in"""
                conditions = []
                params = []

                if from_var.get().strip():
                    conditions.append("date >= %s")
                    params.append(from_var.get().strip())

                if to_var.get().strip():
                    conditions.append("date <= %s")
                    params.append(to_var.get().strip())

                if conditions:
                    query += " WHERE " + " AND ".join(conditions)

                query += " ORDER BY supplier_name"
                cursor.execute(query,params)
                records = cursor.fetchall()

                for row in records:
                    tree.insert("","end",values=(
                        row["supplier_name"] or "",
                        row["phone"] or "",
                        row["email"] or "",
                        row["supplier_cnic"] or "",
                        row["company_name"] or ""
                    ))

                records_label.config(text=f"Total Records: {len(records)}")
                amount_label.config(text="Total Amount: N/A")

            except Exception as e:
                messagebox.showerror("Report Error",str(e))
            finally:
                if conn:
                    conn.close()

        # ========== CUSTOMER REPORT ==========
        elif report == "Customer Report":
            columns = ("Customer","Phone","Email","CNIC","Company")
            tree["columns"] = columns

            for col in columns:
                tree.heading(col,text=col)
                tree.column(col,width=150,minwidth=80,anchor="center")

            conn = None
            try:
                conn = get_connection()
                cursor = conn.cursor()

                query = """SELECT DISTINCT customer_name,phone,email,
                        customer_cnic,company_name FROM stock_out"""
                conditions = []
                params = []

                if from_var.get().strip():
                    conditions.append("date >= %s")
                    params.append(from_var.get().strip())

                if to_var.get().strip():
                    conditions.append("date <= %s")
                    params.append(to_var.get().strip())

                if conditions:
                    query += " WHERE " + " AND ".join(conditions)

                query += " ORDER BY customer_name"
                cursor.execute(query,params)
                records = cursor.fetchall()

                for row in records:
                    tree.insert("","end",values=(
                        row["customer_name"] or "",
                        row["phone"] or "",
                        row["email"] or "",
                        row["customer_cnic"] or "",
                        row["company_name"] or ""
                    ))

                records_label.config(text=f"Total Records: {len(records)}")
                amount_label.config(text="Total Amount: N/A")

            except Exception as e:
                messagebox.showerror("Report Error",str(e))
            finally:
                if conn:
                    conn.close()

                # ========== ✅ CURRENT STOCK REPORT ==========
        elif report == "Current Stock Report":
            columns = ("Thread", "Size", "Available Qty")
            tree["columns"] = columns

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=180, minwidth=120, anchor="center")

            conn = None
            try:
                from database import get_available_stock
                
                conn = get_connection()
                cursor = conn.cursor()

                # Get all unique thread + size combinations from stock_in
                cursor.execute("""
                    SELECT DISTINCT thread_name, size 
                    FROM stock_in 
                    WHERE thread_name IS NOT NULL AND size IS NOT NULL
                """)
                products = cursor.fetchall()

                total_available = 0

                for product in products:
                    thread = product['thread_name']
                    size = product['size']

                    # Calculate total available stock
                    cursor.execute("""
                        SELECT id FROM stock_in
                        WHERE thread_name = %s AND size = %s
                    """, (thread, size))
                    stock_ids = cursor.fetchall()
                    
                    available = 0
                    for row in stock_ids:
                        available += get_available_stock(row['id'])
                    
                    total_available += available

                    tree.insert("", "end", values=(
                        thread,
                        size,
                        available
                    ))

                records_label.config(text=f"Total Products: {len(products)}")
                amount_label.config(text=f"Total Available Stock: {total_available} Bundles")

            except Exception as e:
                messagebox.showerror("Report Error", str(e))
            finally:
                if conn:
                    conn.close()

        # ========== ✅ DYEING STOCK REPORT (NEW) ==========
        elif report == "Dyeing Stock Report":
            columns = ("Color", "Thread", "Size", "Returned Qty", "Sold Qty", 
                      "Available (Not Sold)")
            tree["columns"] = columns

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=130, minwidth=90, anchor="center")

            conn = None
            try:
                conn = get_connection()
                cursor = conn.cursor()

                # Get all returned dyeing records with color
                cursor.execute("""
                    SELECT DISTINCT color, thread_name, size
                    FROM return_dyeing
                    WHERE color IS NOT NULL AND color != ''
                    ORDER BY color, thread_name
                """)
                products = cursor.fetchall()

                total_available = 0

                for product in products:
                    color = product['color']
                    thread = product['thread_name']
                    size = product['size']

                    # Total Returned from Dyeing (for this color)
                    cursor.execute("""
                        SELECT COALESCE(SUM(return_quantity), 0) as total
                        FROM return_dyeing
                        WHERE color = %s AND thread_name = %s AND size = %s
                    """, (color, thread, size))
                    returned_qty = cursor.fetchone()['total'] or 0

                    # Sold (Stock Out) with this color
                    cursor.execute("""
                        SELECT COALESCE(SUM(bundle_quantity), 0) as total
                        FROM stock_out
                        WHERE color = %s AND thread_name = %s AND size = %s
                    """, (color, thread, size))
                    sold_qty = cursor.fetchone()['total'] or 0

                    # Available (Returned but not sold)
                    available = returned_qty - sold_qty
                    total_available += available

                    if available > 0:  # Only show if available
                        tree.insert("", "end", values=(
                            color,
                            thread or "N/A",
                            size or "N/A",
                            returned_qty,
                            sold_qty,
                            available
                        ))

                records_label.config(text=f"Total Color Records: {len(tree.get_children())}")
                amount_label.config(text=f"Total Available (Not Sold): {total_available} Bundles")

                if len(tree.get_children()) == 0:
                    messagebox.showinfo("No Data", "No dyeing stock available (all returned stock has been sold).")

            except Exception as e:
                messagebox.showerror("Report Error", str(e))
            finally:
                if conn:
                    conn.close()

        else:
            messagebox.showinfo("Coming Soon", f"{report} will be added next.")

    # ============================================================
    # EXPORT FUNCTIONS (PDF & EXCEL)
    # ============================================================
    def export_pdf():
        if not tree.get_children():
            messagebox.showwarning("No Data", "Please generate a report first.")
            return

        filename = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF File", "*.pdf")])
        if not filename:
            return

        try:
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)

            styles = getSampleStyleSheet()
            company_style = ParagraphStyle("Company",parent=styles["Normal"],fontName="Helvetica-Bold",
                fontSize=18,textColor=colors.HexColor("#0D3373"),alignment=0,spaceAfter=3)

            subtitle_style = ParagraphStyle("Subtitle",parent=styles["Normal"],
                fontName="Helvetica",fontSize=9,alignment=1,spaceAfter=3)

            title_style = ParagraphStyle("Title",parent=styles["Normal"],fontName="Helvetica-Bold",
                fontSize=15,textColor=colors.HexColor("#0D3373"),alignment=1,spaceBefore=5,spaceAfter=12)

            elements = []
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),"images/company_logo.jpeg")

            if os.path.exists(logo_path):
                from reportlab.platypus import Image
                img = Image(logo_path, width=55, height=55)

                header_data = [[img,
                    Paragraph("<b>RASHID BROTHERS</b><br/>""<font name='Helvetica' size='8'>" \
                    "Manufacturer Of Leather & Leather Goods</font>",company_style),
                    Paragraph("<b>+92-21-35116818</b><br/>""rashidbrothers371@gmail.com<br/>"
                        "Karachi, Pakistan</font>",subtitle_style)]]

                header_table = Table(header_data,colWidths=[70, 430, 220])

                header_table.setStyle(TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (0, 0), (0, 0), "LEFT"),
                    ("ALIGN", (1, 0), (1, 0), "CENTER"),
                    ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                    ("LINEBELOW", (0, 0), (-1, 0), 1, colors.black),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8)
                ]))
                elements.append(header_table)

            else:
                elements.append(Paragraph("RASHID BROTHERS", company_style))
                elements.append(Paragraph("Manufacturer Of Leather & Leather Goods",subtitle_style))
            elements.append(Paragraph(report_var.get().upper(),title_style))

            data = [list(tree["columns"])]

            for item in tree.get_children():
                data.append(list(tree.item(item)["values"]))

            table = Table(data, repeatRows=1)

            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                [colors.white, colors.white]),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
                ("TOPPADDING", (0, 0), (-1, 0), 9)
            ]))

            elements.append(table)

            def add_footer(canvas, doc):
                canvas.saveState()
                canvas.setStrokeColor(colors.black)
                canvas.setLineWidth(1)
                canvas.line(doc.leftMargin,45,landscape(A4)[0] - doc.rightMargin,45)

                footer = Paragraph("<b>RASHID BROTHERS</b><br/>""Plot # ST-371 SECTOR 7/A, Korangi Industrial Area, Karachi",subtitle_style)

                footer.wrapOn(canvas, doc.width, 40)
                footer.drawOn(canvas, doc.leftMargin, 10)

                canvas.restoreState()

            doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
            messagebox.showinfo("Success","PDF report generated successfully! ✅")

            os.startfile(filename)
        except Exception as e:
            messagebox.showerror("PDF Error", str(e))

    def export_excel():
        if not tree.get_children():
            messagebox.showwarning("No Data", "Please generate a report first.")
            return

        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel File", "*.xlsx")])
        if not filename:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            ws.merge_cells("A1:L1")
            ws["A1"] = "RASHID BROTHERS"
            ws["A1"].font = Font(size=20, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:L2")
            ws["A2"] = "Manufacturer Of Leather & Leather Goods"
            ws["A2"].font = Font(size=11)
            ws["A2"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A3:L3")
            ws["A3"] = "+92-21-35116818 | rashidbrothers371@gmail.com | Karachi, Pakistan"
            ws["A3"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A4:L4")
            ws["A4"] = report_var.get().upper()
            ws["A4"].font = Font(size=15, bold=True)
            ws["A4"].alignment = Alignment(horizontal="center")

            columns = tree["columns"]

            for col_num, col in enumerate(columns, start=1):
                cell = ws.cell(row=6, column=col_num, value=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row_num, item in enumerate(tree.get_children(), start=7):
                values = tree.item(item)["values"]
               
                for col_num, value in enumerate(values, start=1):
                    ws.cell(row=row_num,column=col_num,value=value)

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value:
                        max_length = max(max_length,len(str(cell.value)))

                    ws.column_dimensions[column_letter].width = min(
                        max_length + 3,30)
                wb.save(filename)
                messagebox.showinfo("Success","Excel report generated successfully! ✅")

                os.startfile(filename)

        except Exception as e:
            messagebox.showerror("Excel Error", str(e))

    # ============================================================
    # BUTTONS
    # ============================================================
    tk.Button(button_frame,text="🔍\nView Report",command=load_report,font=("Segoe UI", 10, "bold"),
                bg="#27ae60",fg="white",width=12,height=2,relief="flat",cursor="hand2").pack(side="left", padx=8)

    tk.Button(button_frame,text="🔄\nRefresh",command=load_report,font=("Segoe UI", 10, "bold"),
        bg="#2980b9",fg="white",width=12,height=2,relief="flat",cursor="hand2").pack(side="left", padx=8)

    tk.Button(button_frame,text="📄\nPDF",command=export_pdf,font=("Segoe UI", 10, "bold"),bg="#c0392b",
        fg="white",width=12,height=2,relief="flat",cursor="hand2").pack(side="left", padx=8)

    tk.Button(button_frame,text="📊\nExcel",command=export_excel,font=("Segoe UI", 10, "bold"),bg="#16a085",
        fg="white",width=12,height=2,relief="flat",cursor="hand2").pack(side="left", padx=8)

    # Initial report
    load_report()
